from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from selenium.webdriver.support import expected_conditions as EC
from flask import Flask, send_file, render_template, request
from selenium.webdriver.support.wait import WebDriverWait 
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from openpyxl.styles.numbers import FORMAT_NUMBER_00
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from selenium.webdriver.common.by import By
from selenium import webdriver
import sys

def parser(url):
    s=Service(ChromeDriverManager().install())
    option = webdriver.ChromeOptions()
    option.add_argument('headless')
    option.add_argument('--no-sandbox')
    driver = webdriver.Chrome(service=s,options=option)
    driver.get(url)

    print('web page loaded')
    try:
        wb = load_workbook('template.xlsx')
        ws = wb.active

        # parsing stuff

        #name = WebDriverWait(driver, 10).until(
        #    EC.presence_of_element_located((By.CLASS_NAME, 'ui.large.header'))
        #)
        efFoils = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CLASS_NAME, 'effoIL'))
        )
        number = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'hDCZbQ'))
        )
        
        number = (number.text).split('Котировочная сессия')[1]
        
        #date = (efFoils[3].text).split('по')[1]
        orderer = efFoils[1].text.split('Заказчик')[1].split('\n')[1]

        info = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CLASS_NAME, 'gbkFTI'))
        )
        for i in info:
            i.click()
        itemDiv = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CLASS_NAME, 'ecxxjK'))
        )
        headers = []
        manufacturers = []
        model = []
        amount = []
        measurement = []
        cost = []
        total = []
        for item in itemDiv:
            headers.append(
                item.find_element(By.TAG_NAME,'a').get_attribute('textContent')
            )
            
            listOfInfo = item.find_elements(By.CLASS_NAME,'effoIL')
            manufacturers.append(''.join(listOfInfo[9-1].get_attribute('textContent').split('Производитель')))
            model.append(''.join(listOfInfo[8-1].get_attribute('textContent').split('Модель')))
            amount.append((''.join(listOfInfo[1-1].get_attribute('textContent').split('Количество'))).split(' ')[0])
            measurement.append((''.join(listOfInfo[1-1].get_attribute('textContent').split('Количество'))).split(' ')[1])
            costVar = (listOfInfo[2-1].find_element(By.TAG_NAME,'div')).get_attribute('textContent').replace(u'\xa0', u'')
            cost.append(float(costVar.replace(costVar[len(costVar)-1],'').replace(',','.')))
            totalVar =(listOfInfo[3-1].find_element(By.TAG_NAME,'div')).get_attribute('textContent').replace(u'\xa0', u'')
            total.append(float(totalVar.replace(costVar[len(costVar)-1],'').replace(',','.')))

        print('data scraped')

        #styles for table
        font = Font(
            name='Times New Roman',
            bold = False,
            size = 12
        )

        footerFont = Font(
            name = 'Times New Roman',
            bold = True,
            size = 12
        )

        boldness = Font(
            bold = True,
        )

        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))

        profit = Font(
            bold=True,
            size=14
        )

        yellowish = PatternFill(start_color="ffda65", end_color="ffda65", fill_type = "solid")
        greenish = PatternFill(start_color="e2efd9", end_color="e2efd9", fill_type = "solid")

        centralised = Alignment(
            horizontal='center'
        )

        #putting info from header of web page into the header of the table
        ws['C1'].value = number
        ws['C2'].value = url
        ws['C3'].value = orderer

        info = [headers,manufacturers,model,measurement,amount,cost,total ]

        #putting info about items from web page into the table
        for row in range(6,len(itemDiv)+6):
            ws['A'+str(row)].value = row-5
            ws['A'+str(row)].border = thin_border
            for column in range(2,len(info)+2):
                col = get_column_letter(column)
                ws[col+str(row)].value = info[column-2][row-6]
                ws[col+str(row)].border = thin_border
                ws[col+str(row)].font = font
                if column >= 5:
                    ws[col+str(row)].alignment = centralised
                    ws[col+str(row)].number_format = FORMAT_NUMBER_00

        for column in range(1,24):
            col = get_column_letter(column)
            ws[col + str(6+len(itemDiv))].border = thin_border
            ws[col + str(6+len(itemDiv))].font = boldness

        #clearing unnecessary stuff
        for row in range(8,len(itemDiv)+7):
            for column in range(9,24):
                col = get_column_letter(column)
                if column == 18:
                    ws[col+str(row)].fill = yellowish
                if column in [16,15,12,11]:
                    ws[col+str(row)].fill = greenish
                ws[col+str(row)].value = ''
                ws[col+str(row)].border = thin_border
                ws[col+str(row)].font = font

        ws['L19'].value = ''
        ws['L20'].value = ''

        ws['N19'].style = 'Normal'
        ws['N20'].style = 'Normal'
        ws['L19'].style = 'Normal'
        ws['L20'].style = 'Normal'
        ws['V15'].style = 'Normal'
        ws['V17'].style = 'Normal'
        ws['V17'].border = thin_border
        ws['L19'].border = thin_border
        ws['L20'].border = thin_border
        ws['N19'].border = thin_border
        ws['N20'].border = thin_border
        ws['L19'].fill = greenish
        ws['L20'].fill = greenish

        #under yellow segment
        ws['B'+str(5+len(itemDiv)+1)].value = 'ВСЕГО'
        ws['H'+str(5+len(itemDiv)+1)].value = '=SUM(H'+ str(6) +':H'+ str(5+len(itemDiv)) +')'

        #under green segment
        ws['L'+str(13+len(itemDiv))].value = 'Срок поставки'
        ws['L'+str(14+len(itemDiv))].value = 'Количество адресов'

        ws['L'+str(13+len(itemDiv))].font = boldness
        ws['L'+str(14+len(itemDiv))].font = boldness

        ws['P'+str(5+len(itemDiv)+1)].value = '=SUM(P'+ str(6) +':P'+ str(5+len(itemDiv)) +')'
        #under blue segment
        ws['M'+str(5+len(itemDiv)+2)].value = 'К-т наценки'
        ws['M'+str(5+len(itemDiv)+3)].value = 'С наценкой'
        ws['M'+str(5+len(itemDiv)+4)].value = 'Доставка'
        ws['M'+str(5+len(itemDiv)+5)].value = 'ИТОГО'

        ws['N'+str(5+len(itemDiv)+1)].value = '=SUM(N'+ str(6) +':N'+ str(5+len(itemDiv)) +')'
        ws['N'+str(5+len(itemDiv)+2)].value = '1,1'
        ws['N'+str(5+len(itemDiv)+3)].value = '=N'+ str(5+len(itemDiv)+1) +'*N'+ str(5+len(itemDiv)+2)
        ws['N'+str(5+len(itemDiv)+4)].value = '1000'
        ws['N'+str(5+len(itemDiv)+5)].value = '=N'+ str(5+len(itemDiv)+3) +'+N'+ str(5+len(itemDiv)+4)

        ws['M'+str(5+len(itemDiv)+5)].font = boldness
        ws['N'+str(5+len(itemDiv)+1)].font = boldness
        ws['N'+str(5+len(itemDiv)+5)].font = boldness

        #under red segment
        ws['V'+str(5+len(itemDiv)+1)].value = '=SUM(V'+ str(6) +':V'+ str(5+len(itemDiv)) +')'
        ws['T'+str(9+len(itemDiv))].value = 'Выиграли по цене'
        ws['T'+str(11+len(itemDiv))].value = 'Прибыль'

        ws['V'+str(5+len(itemDiv)+1)].font = footerFont
        ws['T'+str(9+len(itemDiv))].font = boldness
        ws['T'+str(11+len(itemDiv))].font = profit

        ws['V'+str(9+len(itemDiv))].border = thin_border
        ws['V'+str(11+len(itemDiv))].border = thin_border

        ws['N'+str(13+len(itemDiv))].border = thin_border
        ws['N'+str(14+len(itemDiv))].border = thin_border
        #saving table as parsed.xlsx
        wb.save(str(number) + '.xlsx')
        print('table saved')
        driver.quit()
        return True,number 
    except:
        print(sys.exc_info()[0])
        driver.quit()
        return False,0

app = Flask(__name__)

@app.route('/', methods = ['GET','POST'])
def hello_world():
    error = ''
    if request.method == 'POST':
        data = request.form.get('link')
        if 'zakupki.mos.ru' in data:
            starting, id = parser(data)
            if starting:
                error = 'всё хорошо. Таблица с информацией скачивается'
                p = str(id) + '.xlsx'
                return send_file(p,as_attachment=True)
            else:
                error = 'что-то пошло не так'
        else:
            error = 'не тот сайт'
    
    return render_template('index.html', error=error)

@app.route('/download')
def download_file():
    p = 'parsed.xlsx'
    return send_file(p,as_attachment=True)

if __name__ == '__main__':
    app.run(debug = True)